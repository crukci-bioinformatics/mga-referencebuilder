#!/usr/bin/env python3

import sys
import time
import urllib.request
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

GEBA_ACC_COL = 24
GEBA_IMG_ID_COL = 0
GEBA_GENOME_NAME_COL = 2
GEBA_GENOME_SIZE_COL = 8

NCBI_ACCESSION_COL = 0
NCBI_WGS_COL = 3
NCBI_REFSEQ_COL = 4
NCBI_NAME_COL = 7
NCBI_ASSEMBLY_COL = 11
NCBI_SUBMITTER_COL = 16
NCBI_URL_COL = 19

REFER = "reference genome"
REP = "representative genome"
NA = "na"

###############################################################################

def mksciname(name):
  flds = name.split()
  return "%s %s" % (flds[0],flds[1]) if len(flds) > 1 else name

def loadGEBA(fn):
  book = load_workbook(sys.argv[1])
  sheet = book.active
  geba = {}
  count = 0
  for row in sheet.iter_rows(min_row=2):
    name = row[GEBA_GENOME_NAME_COL].value
    sciname = mksciname(name)
    acc = row[GEBA_ACC_COL].value
    size = row[GEBA_GENOME_SIZE_COL].value
    gebaId = row[GEBA_IMG_ID_COL].value
    tup = (sciname,name,acc,size,gebaId)
    if sciname in geba:
      print("WARNING: duplicate species %s in GEBA" % (sciname,))
    else:
      geba[sciname] = tup
      count += 1
  print("Loaded %d rows from GEBA db..." % count)
  return geba

def loadGenomes(fn):
  fd = open(fn)
  genomes = {}
  count = 0
  fd.readline()
  fd.readline()
  for line in fd:
    flds = line.split("\t")
    acc = flds[NCBI_ACCESSION_COL]
    wgs = flds[NCBI_WGS_COL]
    refseq = flds[NCBI_REFSEQ_COL]
    name = flds[NCBI_NAME_COL]
    sciname = mksciname(name)
    asm = flds[NCBI_ASSEMBLY_COL]
    sub = flds[NCBI_SUBMITTER_COL]
    url = flds[NCBI_URL_COL]
#    if refseq == "na":
#      continue
    count += 1
    tup = (acc,wgs,refseq,name,asm,sub,url)
    if sciname in genomes:
      genomes[sciname][refseq].append(tup)
    else:
      genomes[sciname] = {}
      genomes[sciname][REFER] = []
      genomes[sciname][REP] = []
      genomes[sciname][NA] = []
      genomes[sciname][refseq] = [tup]
  print("Loaded %d rows from assembly list..." % (count,))
  return genomes

def chooseMatch(genomes):
  # choose 'Complete Genome' over 'Scaffold' over 'Contig'
  keeper = genomes[0]
  tag = keeper[4]
  for tup in genomes[1:]:
    if tag == 'Complete Genome':
      continue
    elif tag == 'Scaffold' and tup[4] == 'Complete Genome':
      keeper = tup
      tag = tup[4]
    elif tag == 'Contig' and tup[4] != 'Contig':
      keeper = tup
      tag = tup[4]
  return keeper

def chooseRepresentativeSet(geba,genomes):
  refMatch = 0
  repMatch = 0
  naMatch = 0
  noMatch = 0
  results = []
  for k in geba.keys():
#    print(k)
    if k in genomes:
      if len(genomes[k][REFER]) > 0:
        t = chooseMatch(genomes[k][REFER])
        results.append(t)
#        for t in genomes[k][REFER]:
#        print("   %s %s %s %s" % (t[0],t[2],t[3],t[4]))
        refMatch += 1
      elif len(genomes[k][REP]) > 0:
#        for t in genomes[k][REP]:
        t = chooseMatch(genomes[k][REP])
        results.append(t)
#        print("   %s %s %s %s" % (t[0],t[2],t[3],t[4]))
        repMatch += 1
      else:
#        for t in genomes[k][NA]:
        t = chooseMatch(genomes[k][NA])
        results.append(t)
#        print("   %s %s %s %s" % (t[0],t[2],t[3],t[4]))
        naMatch += 1
    else:
      noMatch += 1
  print("%d refMatch   %d repMatch    %d naMatch   %d noMatch" % (refMatch,repMatch,naMatch,noMatch))
  return results

def dumpResults(data,fn):
  fd = open(fn,"w")
  fd.write("Accession\tWGS Master\tRefseq Category\tOrganism\tAssembly Level\tSubmitter\n")
  for flds in data:
    line = "\t".join(flds) + "\n"
    fd.write(line)
  fd.close()

###############################################################################

geba = loadGEBA(sys.argv[1])
genomes = loadGenomes(sys.argv[2])
results = chooseRepresentativeSet(geba,genomes)
dumpResults(results,sys.argv[3])
